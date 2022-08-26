# encoding=utf-8
# 作者：Admin
# 日期：2022/8/8 17:25
# 工具：PyCharm
from io import BytesIO
from openpyxl import Workbook
from flask import Response


def generate_excel(title, data, filename):
    wb = Workbook()
    ws = wb.active

    # 首行列名写入excel
    for i, t in enumerate(title):
        ws.cell(row=1, column=(i + 1)).value = t
    # 数据部分写入excel
    title_fields = [t for t in title]
    for i, _data in enumerate(data):
        one_row = [_data[t] for t in title_fields]
        for j, d in enumerate(one_row):
            ws.cell(row=(i + 2), column=(j + 1)).value = d

    # 传给save函数的不是保存文件名，而是BytesIO流
    sio = BytesIO()
    wb.save(sio)

    response = Response()
    response.headers.add("Content-Type", "application/vnd.ms-excel")
    response.headers.add('Content-Disposition', 'attachment', filename=filename.encode("utf-8").decode("latin1"))
    sio.seek(0)
    response.data = sio.getvalue()
    return response
